#!/usr/bin/env python3

import os
import sys
import pwd
import requests
import pandas as pd
import inspect
import argparse
import json
import locale
import re
import zipfile

from configparser import ConfigParser
from tqdm import tqdm
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import OrderedDict


# Setze das Arbeitsverzeichnis auf das Verzeichnis, in dem das Skript gespeichert ist
script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
os.chdir(script_dir)

locale.setlocale(locale.LC_ALL, '')  # Set locale based on system settings

def print_progress(message: str):
    frame = inspect.currentframe().f_back
    filename = os.path.basename(frame.f_code.co_filename)
    line_number = frame.f_lineno
    function_name = frame.f_code.co_name

    progress_message = f"{filename}:{line_number} [{function_name}] {message}"

    if not hasattr(print_progress, "_last_length"):
        print_progress._last_length = 0

    clear_space = max(print_progress._last_length - len(progress_message), 0)
    progress_message += " " * clear_space

    sys.stdout.write(f"\r{progress_message}")
    sys.stdout.flush()

    print_progress._last_length = len(progress_message)

# ---------------------- Configuration Loading ----------------------
def load_config(config_path):
    """Load configuration file."""
    print_progress("process...")
    config = ConfigParser()
    config.read(config_path)
    return config

def get_script_name():
    """Return the name of the current script without extension."""
    return os.path.splitext(os.path.basename(sys.argv[0]))[0]

def load_config_from_script():
    """Load the configuration from the ini file with a priority for the .ufe.ini file."""
    script_name = get_script_name()
    ufe_ini_path = f"{script_name}.ufe.ini"
    ini_path = f"{script_name}.ini"

    # Try to load the .ufe.ini file first
    if os.path.exists(ufe_ini_path):
        print_progress(f"Using config file: {ufe_ini_path}")
        return load_config(ufe_ini_path)
    # Fallback to the .ini file
    elif os.path.exists(ini_path):
        print_progress(f"Using config file: {ini_path}")
        return load_config(ini_path)
    else:
        print(f"Configuration files '{ufe_ini_path}' and '{ini_path}' not found.")
        sys.exit(1)


# ---------------------- Logging ----------------------
def log_message(log_path, message):
    """Append a log message to the log file."""
    with open(log_path, "a") as log_file:
        log_file.write(f"{datetime.now()} - {message}\n")

# ---------------------- API Helpers ----------------------
def fetch_data(url, headers, endpoint):
    """Fetch data from API with pagination."""
    print_progress(message=f"process {endpoint}...")
    data = []
    page = 1
    while True:
        response = requests.get(f"{url}/{endpoint}/?page={page}", headers=headers)

        if response.status_code != 200:
            print(f"Error fetching {endpoint}. HTTP {response.status_code}: {response.text}")
            raise Exception(f"Failed to fetch data from {endpoint}, status code: {response.status_code}")

        try:
            response_data = response.json()
        except requests.exceptions.JSONDecodeError as e:
            print(f"JSON decoding error for {endpoint}: {e}")
            print(f"Response content: {response.text}")
            raise

        data.extend(response_data.get("results", []))
        if not response_data.get("next"):
            break
        page += 1
        print_progress(message=f"process {endpoint}/{page}...")
    return data

def get_name_from_id(url, headers, endpoint, id):
    """Get name from ID using API."""
    response = requests.get(f"{url}/{endpoint}/{id}/", headers=headers)
    if response.status_code == 200:
        return response.json().get("name", "")
    return "Unknown"

# ---------------------- Document Export Helpers ----------------------
def export_pdf(doc_id, doc_title, tag_directory, url, headers):
    """Export a document's PDF."""
    sanitized_title = sanitize_filename(doc_title)
    pdf_path = os.path.join(tag_directory, f"{sanitized_title}.pdf")
    response = requests.get(f"{url}/documents/{doc_id}/download/", headers=headers)
    if response.status_code == 200:
        with open(pdf_path, "wb") as pdf_file:
            pdf_file.write(response.content)
    else:
        print(f"Failed to download PDF for {doc_title}")

def sanitize_filename(filename):
    """
    Remove or replace characters in the filename that are not allowed in file names.
    """
    sanitized = re.sub(r'[<>:"/\\|?*]', '-', filename)  # Ersetze verbotene Zeichen durch '-'
    return sanitized[:255]  # Truncate to avoid overly long filenames

def export_json(doc_data, doc_title, tag_directory):
    """Export a document's metadata as JSON."""
    sanitized_title = sanitize_filename(doc_title)
    json_path = os.path.join(tag_directory, f"{sanitized_title}.json")
    with open(json_path, "w", encoding="utf-8") as json_file:
        json.dump(doc_data, json_file, ensure_ascii=False, indent=4)

# ---------------------- Excel Export Helpers ----------------------
def export_to_excel(data, file_path, script_name, tag_name, api_url, custom_fields_map,currency_columns):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import os
    import pwd

    # API-Basis-URL ohne `/api` generieren
    base_url = api_url.rstrip("/api")

    # Ordnerpfad aus file_path extrahieren
    directory = os.path.dirname(file_path)

    # Dateiname vorbereiten
    base_filename = f"##export-{tag_name}-{datetime.now().strftime('%Y%m%d')}"
    file_ext = ".xlsx"
    filename = f"{base_filename}{file_ext}"
    fullfilename = os.path.join(directory, filename)

    # Falls Datei bereits geöffnet oder existiert, iterativ neuen Namen finden
    counter = 1
    while os.path.exists(fullfilename):
        filename = f"{base_filename}-{counter}{file_ext}"
        fullfilename = os.path.join(directory, filename)
        counter += 1

    # Pandas DataFrame aus document_data erstellen
    df = pd.DataFrame(data)


    with pd.ExcelWriter(fullfilename, engine="openpyxl") as writer:
        # DataFrame in Excel schreiben (ab Zeile 3 für Daten)
        df.to_excel(writer, index=False, startrow=2, sheet_name="Dokumentenliste")
        worksheet = writer.sheets["Dokumentenliste"]

        # Headerzeile (A1) mit Scriptnamen, Tag und anderen Infos
        header_info = f"{script_name} -- {tag_name} -- {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} -- {pwd.getpwuid(os.getuid()).pw_name} -- {os.uname().nodename}"
        worksheet["A1"] = header_info
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))  # Header über alle Spalten
        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")  # Dunkelblau
        worksheet["A1"].font = header_font
        worksheet["A1"].fill = header_fill

        # Summenzeilen für Currency-Spalten in Zeile 2
        for column_name in currency_columns:
            if column_name in df.columns:
                col_idx = df.columns.get_loc(column_name) + 1  # Excel-Spaltenindex
                start_cell = worksheet.cell(row=4, column=col_idx).coordinate
                end_cell = worksheet.cell(row=worksheet.max_row, column=col_idx).coordinate
                sum_formula = f"=SUM({start_cell}:{end_cell})"
                sum_cell = worksheet.cell(row=2, column=col_idx)
                sum_cell.value = sum_formula
                sum_cell.font = Font(bold=True)

        # Spaltentitel (Zeile 3)
        header_row = worksheet[3]
        for cell in header_row:
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # Autofilter
        worksheet.auto_filter.ref = f"A3:{worksheet.cell(row=3, column=len(df.columns)).coordinate}"

        # Definiere die Formate für gerade und ungerade Zeilen
        light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        font = Font(name="Arial", size=11)

        # Formeln für gerade und ungerade Zeilen
        formula_even = "MOD(ROW(),2)=0"
        formula_odd = "MOD(ROW(),2)<>0"

        # Bereich, der formatiert werden soll
        range_string = f"A4:{worksheet.cell(row=worksheet.max_row, column=len(df.columns)).coordinate}"

        # Bedingte Formatierung für gerade Zeilen
        rule_even = FormulaRule(formula=[formula_even], fill=light_blue_fill, font=font)
        worksheet.conditional_formatting.add(range_string, rule_even)

        # Bedingte Formatierung für ungerade Zeilen
        rule_odd = FormulaRule(formula=[formula_odd], fill=white_fill, font=font)
        worksheet.conditional_formatting.add(range_string, rule_odd)

        # Hyperlinks in der ID-Spalte
        # Suche die Spalte basierend auf dem Header in Zeile 3
        document_column = "ID"  # Der Header-Name für die Spalte mit den Dokument-IDs
        id_column_idx = None
        for col_idx, cell in enumerate(worksheet[3], start=1):  # Zeile 3 ist der Header
            if cell.value == document_column:
                id_column_idx = col_idx
                break

        # Dokument-ID in URLs umwandeln
        if id_column_idx:  # Wenn die Spalte mit der ID gefunden wurde
            for row_idx in range(4, worksheet.max_row + 1):  # Daten beginnen in Zeile 4
                doc_id = worksheet.cell(row=row_idx, column=id_column_idx).value
                if doc_id:  # Nur wenn ein Wert vorhanden ist
                    link_formula = f'=HYPERLINK("{base_url}/documents/{doc_id}/details", "{doc_id}")'
                    worksheet.cell(row=row_idx, column=id_column_idx).value = link_formula

        # Schriftart-Objekt definieren
        default_font = Font(name="Arial")

        # Alle Zellen formatieren
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.font = default_font


    print(f"\nExcel-Datei erfolgreich erstellt: {fullfilename}")
 
def get_custom_field_definitions(url, headers):
    """Holt die Definitionen für Custom Fields und erzeugt Mappings für Namen, Typ und Auswahloptionen."""
    print_progress(message=f"process custom fields...")
    custom_fields_response = requests.get(f"{url}/custom_fields/", headers=headers)
    custom_fields_map = {}
    custom_field_choices_map = {}

    if custom_fields_response.status_code == 200:
        try:
            custom_fields_data = custom_fields_response.json()
            for field in custom_fields_data["results"]:
                field_id = field["id"]
                field_name = field["name"]
                print_progress(message=f"process {field_name}...")

                field_type = field["data_type"]

                # Speichern der Felddefinitionen im custom_fields_map
                custom_fields_map[field_id] = {
                    "name": field_name,
                    "type": field_type,
                    "choices": {}
                }

                # Wenn es Auswahloptionen gibt, speichern wir sie
                if field_type == "select":
                    custom_field_choices_map[field_id] = {
                        idx: option for idx, option in enumerate(field["extra_data"]["select_options"])
                    }
                    # Füge die Optionen dem custom_fields_map hinzu
                    custom_fields_map[field_id]["choices"] = custom_field_choices_map[field_id]
        except json.decoder.JSONDecodeError as e:
            print(f"JSON-Dekodierungsfehler beim Abrufen der Custom Fields: {e}")
            exit()
    else:
        print(f"Fehler beim Abrufen der Custom Fields. Status Code: {custom_fields_response.status_code}")
        exit()

    return custom_fields_map


def has_file_from_today(directory):
    """
    Prüft, ob im angegebenen Verzeichnis eine Datei existiert,
    die heute erstellt oder zuletzt geändert wurde.
    """
    today = datetime.now().date()
    if not os.path.exists(directory):
        return False

    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if os.path.isfile(file_path):
            # Änderungszeitpunkt der Datei
            file_mtime = datetime.fromtimestamp(os.path.getmtime(file_path))
            if file_mtime.date() == today:
                return True
    return False

# ---------------------- Main Export Logic ----------------------


def process_documents_by_tag(documents, tag_name, tag_id, url, headers, custom_fields_map, export_directory, log_file, tag_dict, script_name, is_all_docs=False):
    """Process and export documents by tag or all documents."""
    tag_dir = os.path.join(export_directory, f"{tag_name}")
      
    # Wenn is_all_docs True, prüfe, ob bereits heute ein Export durchgeführt wurde
    if is_all_docs and has_file_from_today(tag_dir):
        log_message(log_path=log_file,message=f"Export für alle Dokumente wurde übersprungen, da bereits heute Dateien exportiert wurden.")
        return
    os.makedirs(tag_dir, exist_ok=True)

    document_data = []
    currency_columns = []  # Liste zur Speicherung aller Currency-Felder
    custom_fields = {}

    for doc in tqdm(documents, desc=f"Processing documents for tag '{tag_name}'", unit="doc"):
        # Wenn 'is_all_docs' True ist, exportiere alle Dokumente, unabhängig vom Tag
        if not is_all_docs and tag_id not in doc.get("tags", []):
            continue

        detailed_response = requests.get(f"{url}/documents/{doc['id']}/", headers=headers)
        detailed_doc = detailed_response.json()

        custom_fields, doc_currency_columns = process_custom_fields(custom_fields_map, detailed_doc)
        currency_columns.extend(doc_currency_columns)  # Speichere Currency-Felder

       # Dokumentdaten sammeln
        row = OrderedDict([
            ("ID", doc.get("id")),
            ("AddDateFull", format_date(parse_date(doc.get("added")), "yyyy-mm-dd")),
            ("Korrespondent", get_name_from_id(url, headers, "correspondents", doc.get("correspondent"))),
            ("Titel", doc.get("title")),
            ("Tags", ", ".join(tag_dict.get(tag_id, f"Tag {tag_id}") for tag_id in doc.get("tags", []))),

            # Custom Fields direkt hinter den Tags einfügen
            *custom_fields.items(),  

            ("ArchivDate", parse_date(doc.get("created"))),
            ("ArchivedDateMonth", format_date(parse_date(doc.get("created")), "yyyy-mm")),
            ("ArchivedDateFull", format_date(parse_date(doc.get("created")), "yyyy-mm-dd")),

            ("ModifyDate", parse_date(doc.get("modified"))),
            ("ModifyDateMonth", format_date(parse_date(doc.get("modified")), "yyyy-mm")),
            ("ModifyDateFull", format_date(parse_date(doc.get("modified")), "yyyy-mm-dd")),

            ("AddedDate", parse_date(doc.get("added"))),
            ("AddDateMonth", format_date(parse_date(doc.get("added")), "yyyy-mm")),
            ("AddDateFull", format_date(parse_date(doc.get("added")), "yyyy-mm-dd")),

            ("Seiten", doc.get("page_count")),
            ("Dokumenttyp", get_name_from_id(url, headers, "document_types", doc.get("document_type"))),
            ("Speicherpfad", get_name_from_id(url, headers, "storage_paths", doc.get("storage_path"))),
            ("OriginalName", doc.get("original_file_name")),
            ("ArchivedName", doc.get("archived_file_name")),
            ("Owner", doc.get("Owner")),
            ("Notes", doc.get("Notes")),
        ])

        document_data.append(row) 

        export_pdf(doc['id'], doc['title'], tag_dir, url, headers)
        export_json(detailed_doc, doc['title'], tag_dir)
        document_data.append(row)

    excel_file = os.path.join(tag_dir, f"##{tag_name}-{datetime.now().strftime('%Y%m%d')}.xlsx")
    export_to_excel(document_data, excel_file, script_name, tag_name, api_url=url, custom_fields_map=custom_fields_map, currency_columns=currency_columns)
    log_message(log_file, f"Tag: {tag_name}, Documents exported: {len(document_data)}")
    print(f"Exported Excel file: {excel_file}")

def process_custom_fields(custom_fields_map, detailed_doc):
    custom_fields = {}
    currency_fields = []  # Liste zum Speichern der Currency-Feldnamen

    if "custom_fields" in detailed_doc:
        for custom_field in detailed_doc.get("custom_fields", []):
            field_id = custom_field.get("field")
            if not field_id:    
                continue

            field_info = custom_fields_map.get(field_id, {})
            if not isinstance(field_info, dict):
                continue

            field_name = field_info.get("name", f"Feld {field_id}")
            field_type = field_info.get("type", "string")
            field_value = custom_field.get("value", "")

            if field_type == "monetary":
                numeric_value = parse_currency(field_value)
                custom_fields[field_name] = numeric_value  # Rohdaten speichern
                custom_fields[f"{field_name}_formatted"] = format_currency(field_value)  # Formatierte Version speichern
                currency_fields.append(field_name)  # Speichern des Currency-Felds
            elif field_type == "select":
                if isinstance(field_info.get("choices"), dict):
                    custom_fields[field_name] = field_info["choices"].get(field_value, f"Wert {field_value}")
                else:
                    custom_fields[field_name] = f"Wert {field_value}"
            else:
                custom_fields[field_name] = field_value

    return custom_fields, currency_fields

def parse_currency(value):
    """Parst einen Währungswert wie 'EUR5.00' in einen Float."""
    try:
        # Entferne Währungszeichen (alles außer Ziffern, Punkt oder Minus)
        numeric_part = ''.join(c for c in value if c.isdigit() or c == '.' or c == '-')
        return float(numeric_part)
    except Exception as e:
        # print(f"Fehler beim Parsen des Währungswerts '{value}': {e}")
        return 0.0  # Fallback auf 0 bei Fehlern

def format_currency(value, currency_locale="de_DE.UTF-8"):
    if value is None:
        return ""
    try:
        clean_value = ''.join(filter(str.isdigit, value))
        if not clean_value:
            return "0,00"
        value_float = float(clean_value) / 100
    except ValueError:
        value_float = 0.0

    locale.setlocale(locale.LC_ALL, currency_locale)
    formatted_value = locale.currency(value_float, grouping=True)
    return formatted_value

from datetime import datetime
from dateutil import parser

def format_date(date_string, output_format):
    """
    Formatiert das Datum im Format '%d.%m.%Y' oder '%d.%m.%Y %H:%M' 
    in das gewünschte Format:
    - 'yyyy-mm' oder
    - 'yyyy-mm-dd'.
    
    Parameter:
    - date_string: Das Datum als String (im Format '%d.%m.%Y' oder '%d.%m.%Y %H:%M').
    - output_format: Das gewünschte Ausgabeformat ('yyyy-mm' oder 'yyyy-mm-dd').
    
    Rückgabe:
    - Das Datum im gewünschten Format als String oder None bei Fehlern.
    """
    if not date_string:
        print(f"Date string is empty or None: {date_string}")
        return None

    try:
        # Datum im ursprünglichen Format parsen
        if len(date_string.split(" ")) > 1:
            parsed_date = datetime.strptime(date_string, "%d.%m.%Y %H:%M")
        else:
            parsed_date = datetime.strptime(date_string, "%d.%m.%Y")
        
        # Rückgabe im gewünschten Format
        if output_format == "yyyy-mm":
            return parsed_date.strftime("%Y-%m")
        elif output_format == "yyyy-mm-dd":
            return parsed_date.strftime("%Y-%m-%d")
        else:
            print(f"Unsupported output format: {output_format}")
            return None
    except Exception as e:
        print(f"Failed to format date '{date_string}': {e}")
        return None

def parse_date(date_string):
    """
    Versucht, das Datum mit oder ohne Zeitzonen-Offset zu parsen.
    Gibt das Datum im Format '%d.%m.%Y' zurück, wenn die Uhrzeit 00:00 ist,
    andernfalls im Format '%d.%m.%Y %H:%M'.
    """
    if not date_string:
        print(f"Date string is empty or None: {date_string}")
        return None

    try:
        # Verwende dateutil.parser, um flexibel zu parsen
        parsed_date = parser.isoparse(date_string)
        
        # Prüfe, ob die Uhrzeit 00:00 ist
        if parsed_date.hour == 0 and parsed_date.minute == 0:
            return parsed_date.strftime("%d.%m.%Y")
        else:
            return parsed_date.strftime("%d.%m.%Y %H:%M")
    except Exception as e:
        print(f"Failed to parse date '{date_string}': {e}")
        return None

def prepare_tag_directory_for_export(tag_name, tag_dir):
    """
    Archiviert alle existierenden Inhalte des Verzeichnisses in einer ZIP-Datei und leert das Verzeichnis.
    
    :param tag_name: Name des Tags
    :param tag_dir: Verzeichnis für den Tag
    """
    # Format der ZIP-Datei: Tagname + Datum
    zip_filename = os.path.join(
        tag_dir,
        f"##{tag_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    )
    
    # ZIP-Datei erstellen
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(tag_dir):
            for file in files:
                if not file.endswith('.zip'):  # Keine existierenden ZIP-Dateien packen
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, tag_dir)  # Relativer Pfad in der ZIP-Datei
                    zipf.write(file_path, arcname)
    
    # Verzeichnis leeren (außer ZIP-Dateien)
    for root, dirs, files in os.walk(tag_dir):
        for file in files:
            if not file.endswith('.zip'):  # ZIP-Dateien behalten
                os.remove(os.path.join(root, file))
        for dir in dirs:
            os.rmdir(os.path.join(root, dir))

def ensure_directory_exists(directory):
    """
    Erstellt das Verzeichnis, falls es nicht existiert.
    """
    if not os.path.exists(directory):
        os.makedirs(directory)

def export_all_documents(tag_name, export_dir, documents, api_url, headers, custom_fields_map, tag_dict, script_name, log_path):
    """
    Exportiert alle Dokumente in ein spezielles Verzeichnis für "ALLDocs".
    """
    all_docs_dir = os.path.join(export_dir, tag_name)
    
    # Verzeichnis sicherstellen
    ensure_directory_exists(all_docs_dir)
    
    # Vor dem Export Inhalte archivieren und Verzeichnis leeren
    prepare_tag_directory_for_export(tag_name, all_docs_dir)
    
    # Export-Prozess starten
    log_message(
        log_path,
        f"Exportiere alle Dokumente in '{tag_name}'"
    )
    process_documents_by_tag(
        documents, tag_name, None, api_url, headers, custom_fields_map,
        tag_dict=tag_dict, export_directory=export_dir,
        script_name=script_name, log_file=log_path
    )


def export_for_tags(tags, export_dir, documents, api_url, headers, custom_fields_map, tag_dict, script_name, log_path):
    """
    Exportiert nur die Tags, für die ein Unterverzeichnis im Export-Verzeichnis existiert.
    Zusätzlich wird ein Export für 'ALLDocs' durchgeführt, falls gewünscht.
    """
    # Export für "ALLDocs"
    all_docs_tag_name = "ALLDocs"
    all_docs_tag_id = -1  # Ein Tag-ID für ALLDocs (optional, da es nur einen Export für alle Dokumente macht)
    
    # Exportiere alle Dokumente, auch ohne spezifisches Tag
    log_message(log_path, f"Exportiere alle Dokumente für 'ALLDocs'")
    process_documents_by_tag(
        documents, all_docs_tag_name, all_docs_tag_id, api_url, headers, custom_fields_map,
        export_directory=export_dir, log_file=log_path, tag_dict=tag_dict, script_name=script_name, is_all_docs=True
    )
    
    # Export für die spezifischen Tags
    for tag in tags:
        tag_name = tag["name"]
        tag_id = tag["id"]
        
        # Pfad zum Unterverzeichnis für das Tag erstellen
        tag_dir = os.path.join(export_dir, tag_name)

        # Überprüfen, ob das Unterverzeichnis existiert
        if os.path.isdir(tag_dir):
            # Vor dem Export: Inhalte archivieren und Verzeichnis leeren
            prepare_tag_directory_for_export(tag_name, tag_dir)
            
            # Dokumente exportieren
            log_message(log_path, f"Exportiere Dokumente für Tag: {tag_name} (ID: {tag_id})")
            process_documents_by_tag(
                documents, tag_name, tag_id, api_url, headers, custom_fields_map,
                tag_dict=tag_dict, export_directory=export_dir,
                script_name=script_name, log_file=log_path
            )
        else:
            # Wenn das Verzeichnis nicht existiert, logge eine Nachricht
            log_message(log_path, f"Verzeichnis für Tag {tag_name} existiert nicht. Export übersprungen.")


# Funktion, um den Log-Dateinamen basierend auf dem Skriptnamen und Datum zu erstellen
def get_log_filename(script_name, log_dir, suffix="progress"):
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if suffix == "log":
        return os.path.join(log_dir, f"##{script_name}__{timestamp}.log")
    else:
        return os.path.join(log_dir, f"##{script_name}__{timestamp}.{suffix}.log")

# Funktion, um Logs zu initialisieren
def initialize_log(log_dir, script_name):
    final_log_path = get_log_filename(script_name, log_dir, "log")
    progress_log_path = get_log_filename(script_name, log_dir, "progress")
    
    # Falls ein vorheriges Log existiert, es in die neue Log-Datei kopieren
    if os.path.exists(final_log_path):
        with open(progress_log_path, "w") as new_log, open(final_log_path, "r") as old_log:
            shutil.copyfileobj(old_log, new_log)
        os.remove(final_log_path)
    else:
        open(progress_log_path, "w").close()  # Erstelle eine leere Log-Datei
    
    return progress_log_path, final_log_path

# Funktion, um das Log umzubenennen
def finalize_log(progress_log_path, final_log_path):
    if os.path.exists(progress_log_path):
        os.rename(progress_log_path, final_log_path)

# ---------------------- Main ----------------------
def main():
    script_name = get_script_name()
    config = load_config_from_script()

    export_dir = config.get("Export", "directory")
    api_url = config.get("API", "url")
    api_token = config.get("API", "token")
    log_dir = config.get("Log", "log_file")

    headers = {"Authorization": f"Token {api_token}"}

    # Log-Dateien initialisieren
    progress_log_path, final_log_path = initialize_log(log_dir, script_name)
    
    # Logging starten
    log_message(progress_log_path, "Starting export...")

    try:
        tags = fetch_data(api_url, headers, "tags")
        documents = fetch_data(api_url, headers, "documents")
        custom_fields_map = get_custom_field_definitions(url=api_url, headers=headers)
        tag_dict = {tag["id"]: tag["name"] for tag in tags}

        # Neue Funktion aufrufen, die den Export durchführt
        export_for_tags(
            tags,
            export_dir=export_dir,
            documents=documents,
            api_url=api_url,
            headers=headers,
            custom_fields_map=custom_fields_map,
            tag_dict=tag_dict,
            script_name=script_name,
            log_path=progress_log_path,
        )
    except Exception as e:
        log_message(progress_log_path, f"Error: {str(e)}")
        raise
    finally:
        # Log umbenennen
        finalize_log(progress_log_path, final_log_path)

if __name__ == "__main__":
    main()

