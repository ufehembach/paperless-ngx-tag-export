#!/usr/bin/env python3

import os
import sys
import requests
import pandas as pd
import argparse
from configparser import ConfigParser
from tqdm import tqdm
import json
from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
import locale

locale.setlocale(locale.LC_ALL, '')  # Set locale based on system settings

# ---------------------- Configuration Loading ----------------------
def load_config(config_path):
    """Load configuration file."""
    config = ConfigParser()
    config.read(config_path)
    return config

# ---------------------- Logging ----------------------
def log_message(log_path, message):
    """Append a log message to the log file."""
    with open(log_path, "a") as log_file:
        log_file.write(f"{datetime.now()} - {message}\n")

# ---------------------- Progress Printing ----------------------
def print_progress(message: str):
    """Prints progress messages."""
    sys.stdout.write(f"\r{message}")
    sys.stdout.flush()

# ---------------------- API Helpers ----------------------
def fetch_data(url, headers, endpoint):
    """Fetch data from API with pagination."""
    data = []
    page = 1
    while True:
        response = requests.get(f"{url}/{endpoint}?page={page}", headers=headers)
        if response.status_code != 200:
            raise Exception(f"Failed to fetch data from {endpoint}, status code: {response.status_code}")
        response_data = response.json()
        data.extend(response_data.get("results", []))
        if not response_data.get("next"):
            break
        page += 1
    return data

def get_name_from_id(url, headers, endpoint, id):
    """Get name from ID using API."""
    response = requests.get(f"{url}/{endpoint}/{id}/", headers=headers)
    if response.status_code == 200:
        return response.json().get("name", "")
    return "Unknown"

# ---------------------- Document Export Helpers ----------------------
def export_pdf(doc_id, doc_title, tag_directory, url, headers):
    """Export a document's PDF."""
    sanitized_title = sanitize_filename(doc_title)
    pdf_path = os.path.join(tag_directory, f"{sanitized_title}.pdf")
    response = requests.get(f"{url}/documents/{doc_id}/download/", headers=headers)
    if response.status_code == 200:
        with open(pdf_path, "wb") as pdf_file:
            pdf_file.write(response.content)
    else:
        print(f"Failed to download PDF for {doc_title}")

def export_json(doc_data, doc_title, tag_directory):
    """Export a document's metadata as JSON."""
    sanitized_title = sanitize_filename(doc_title)
    json_path = os.path.join(tag_directory, f"{sanitized_title}.json")
    with open(json_path, "w", encoding="utf-8") as json_file:
        json.dump(doc_data, json_file, ensure_ascii=False, indent=4)

# ---------------------- Excel Export Helpers ----------------------
def setup_excel_formatting(writer, sheet_name):
    """Set formatting for the Excel sheet."""
    worksheet = writer.sheets[sheet_name]
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    light_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    formula = "MOD(ROW(),2)=1"
    rule = FormulaRule(formula=[formula], fill=light_blue_fill)
    worksheet.conditional_formatting.add(f"A2:Z{worksheet.max_row}", rule)

def export_to_excel(data, file_path, paperless_url):
    """Export data to an Excel file."""
    df = pd.DataFrame(data)
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Documents")
        setup_excel_formatting(writer, "Documents")
        # Add hyperlink to titles
        worksheet = writer.sheets["Documents"]
        for idx, doc in enumerate(data, start=2):
            link = f"{paperless_url}/documents/{doc['ID']}"
            worksheet[f"B{idx}"].hyperlink = link
            worksheet[f"B{idx}"].style = "Hyperlink"

# ---------------------- Main Export Logic ----------------------
def process_documents_by_tag(documents, tag_name, tag_id, url, headers, custom_fields_map, export_directory, log_file):
    """Process and export documents by tag."""
    tag_dir = os.path.join(export_directory, f"export-{tag_name}")
    os.makedirs(tag_dir, exist_ok=True)

    document_data = []
    for doc in tqdm(documents, desc=f"Processing documents for tag '{tag_name}'", unit="doc"):
        if tag_id not in doc.get("tags", []):
            continue

        detailed_response = requests.get(f"{url}/documents/{doc['id']}/", headers=headers)
        detailed_doc = detailed_response.json()

        row = {
            "ID": doc.get("id"),
            "Title": doc.get("title"),
            "Tags": ", ".join(tag_dict.get(tid, f"Tag {tid}") for tid in doc.get("tags", [])),
            **{cf_map["name"]: detailed_doc["custom_fields"].get(cf_map["id"], "") for cf_map in custom_fields_map.values()}
        }

        export_pdf(doc['id'], doc['title'], tag_dir, url, headers)
        export_json(detailed_doc, doc['title'], tag_dir)
        document_data.append(row)

    excel_file = os.path.join(tag_dir, f"export-{tag_name}-{datetime.now().strftime('%Y%m%d')}.xlsx")
    export_to_excel(document_data, excel_file, url)
    log_message(log_file, f"Tag: {tag_name}, Documents exported: {len(document_data)}")
    print(f"Exported Excel file: {excel_file}")

def process_tags_from_directories(documents, tag_dict, url, headers, custom_fields_map, export_directory, log_file):
    """Process tags dynamically from directories."""
    for tag_name in os.listdir(export_directory):
        tag_path = os.path.join(export_directory, tag_name)
        if not os.path.isdir(tag_path):
            continue

        tag_id = next((tid for tid, tname in tag_dict.items() if tname.lower() == tag_name.lower()), None)
        if not tag_id:
            error_log = os.path.join(tag_path, "tag_not_found.log")
            hostname = os.uname().nodename
            user = os.getlogin()
            with open(error_log, "a") as log:
                log.write(f"{datetime.now()} - Tag '{tag_name}' not found on {hostname} by {user}\n")
            continue

        process_documents_by_tag(documents, tag_name, tag_id, url, headers, custom_fields_map, export_directory, log_file)

# ---------------------- Main Function ----------------------
def main():
    parser = argparse.ArgumentParser(description="Export documents from Paperless-ngx.")
    parser.add_argument("-c", "--config", default="my.ini", help="Path to configuration file")
    args = parser.parse_args()

    config = load_config(args.config)
    url = config['paperless']['url']
    token = config['paperless']['token']
    export_dir = config['paperless']['export_directory']
    headers = {"Authorization": f"Token {token}"}

    log_file = os.path.join(export_dir, "export_log.txt")

    documents = fetch_data(url, headers, "documents")
    tags = fetch_data(url, headers, "tags")

    tag_dict = {tag["id"]: tag["name"] for tag in tags}
    custom_fields_map = get_custom_field_definitions(url, headers)

    process_tags_from_directories(documents, tag_dict, url, headers, custom_fields_map, export_dir, log_file)

if __name__ == "__main__":
    main()
